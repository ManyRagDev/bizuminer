from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/sessions/eager-wonderful-lamport/mnt/outputs/modelo-financeiro.xlsx"

FONT = "Arial"
BLUE = "0000FF"
GREEN = "008000"
BLACK = "000000"
DARK = "1A1A1E"
ACCENT = "E8590C"
YELLOW = PatternFill("solid", fgColor="FFFF00")
HEADFILL = PatternFill("solid", fgColor="1A1A1E")
SOFTFILL = PatternFill("solid", fgColor="F2F2F0")

CUR = 'R$ #,##0.00'
CUR0 = 'R$ #,##0;(R$ #,##0);-'
PCT = '0.0%'
INT = '#,##0'

thin = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=thin)


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, size=15, bold=True, color=DARK)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=FONT, size=10, italic=True, color="666666")


def section(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=ACCENT)


def label(ws, row, text, indent=0):
    c = ws.cell(row=row, column=1, value=("   " * indent) + text)
    c.font = Font(name=FONT, size=10, color=BLACK)
    return c


def inputcell(ws, row, col, value, fmt):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=10, bold=True, color=BLUE)
    c.fill = YELLOW
    c.number_format = fmt
    c.alignment = Alignment(horizontal="right")
    return c


def formulacell(ws, row, col, formula, fmt, bold=False, link=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = Font(name=FONT, size=10, bold=bold, color=(GREEN if link else BLACK))
    c.number_format = fmt
    c.alignment = Alignment(horizontal="right")
    return c


def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, size=9, italic=True, color="808080")
    return c


wb = Workbook()

# ===================== 1. PREMISSAS =====================
ws = wb.active
ws.title = "Premissas"
title(ws, "Premissas do modelo",
      "Edite apenas as células AMARELAS. Todo o resto é calculado a partir delas.")

ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 62

r = 4
section(ws, r, "RECEITA")
r += 1
label(ws, r, "Ticket médio mensal por cliente")
inputcell(ws, r, 2, 97, CUR)
note(ws, r, 3, "Plano Pro como carro-chefe. Premissa não validada.")
r += 1
label(ws, r, "Churn mensal")
inputcell(ws, r, 2, 0.03, PCT)
note(ws, r, 3, "Benchmark de SaaS B2B brasileiro, não do nosso nicho.")

r += 2
section(ws, r, "CUSTOS VARIÁVEIS POR CLIENTE")
r += 1
label(ws, r, "Taxa percentual do gateway")
inputcell(ws, r, 2, 0.039, PCT)
note(ws, r, 3, "Stripe Brasil: 3,9% + R$0,39 por transação bem-sucedida.")
r += 1
label(ws, r, "Taxa fixa do gateway")
inputcell(ws, r, 2, 0.39, CUR)
r += 1
label(ws, r, "Alíquota do Simples COM Fator R (Anexo III)")
inputcell(ws, r, 2, 0.073, PCT)
note(ws, r, 3, "Faixa de R$240 mil/ano. Na 1a faixa (ate R$180 mil/ano) e 6,0%.")
r += 1
label(ws, r, "Alíquota do Simples SEM Fator R (Anexo V)")
inputcell(ws, r, 2, 0.155, PCT)
note(ws, r, 3, "1a faixa do Anexo V. Usada quando a folha fica abaixo de 28% da receita.")
r += 1
label(ws, r, "Infraestrutura variável por cliente")
inputcell(ws, r, 2, 1.50, CUR)
note(ws, r, 3, "Sem sessão de WhatsApp conectada, o custo por cliente é mínimo.")

r += 2
section(ws, r, "CUSTOS FIXOS MENSAIS")
FIXED_START = r + 1
r += 1
label(ws, r, "VPS (aplicação, banco e workers)")
inputcell(ws, r, 2, 150, CUR0)
note(ws, r, 3, "Hetzner CX42 16GB ≈ R$96. Folga incluída.")
r += 1
label(ws, r, "Proxies residenciais")
inputcell(ws, r, 2, 150, CUR0)
note(ws, r, 3, "Uso reduzido: Shopee e Awin são API oficial.")
r += 1
label(ws, r, "Contabilidade")
inputcell(ws, r, 2, 300, CUR0)
r += 1
label(ws, r, "Observabilidade e ferramentas")
inputcell(ws, r, 2, 200, CUR0)
r += 1
label(ws, r, "Domínio e e-mail transacional")
inputcell(ws, r, 2, 50, CUR0)
FIXED_END = r
r += 1
label(ws, r, "Total de custos fixos mensais").font = Font(name=FONT, size=10, bold=True)
formulacell(ws, r, 2, f"=SUM(B{FIXED_START}:B{FIXED_END})", CUR0, bold=True)
FIXED_TOTAL = r

r += 2
section(ws, r, "REMUNERAÇÃO DOS SÓCIOS")
r += 1
label(ws, r, "Pró-labore mensal por sócio")
inputcell(ws, r, 2, 3000, CUR0)
PROL_VAL = r
r += 1
label(ws, r, "Número de sócios com pró-labore")
inputcell(ws, r, 2, 2, INT)
PROL_QTD = r
r += 1
label(ws, r, "Total de pró-labore").font = Font(name=FONT, size=10, bold=True)
formulacell(ws, r, 2, f"=B{PROL_VAL}*B{PROL_QTD}", CUR0, bold=True)
PROL_TOTAL = r
note(ws, r, 3, "O Fator R exige folha ≥28% da receita para manter o Anexo III.")

r += 2
section(ws, r, "AQUISIÇÃO — NOVOS CLIENTES POR MÊS")
r += 1
label(ws, r, "Cenário conservador")
inputcell(ws, r, 2, 5, INT)
CEN_CONS = r
r += 1
label(ws, r, "Cenário base")
inputcell(ws, r, 2, 15, INT)
CEN_BASE = r
r += 1
label(ws, r, "Cenário otimista")
inputcell(ws, r, 2, 35, INT)
CEN_OTIM = r
note(ws, r, 3, "Estimativa. Substituir por dado real após os primeiros 60 dias.")

r += 2
section(ws, r, "PARÂMETROS DE REFERÊNCIA")
r += 1
label(ws, r, "Múltiplo LTV:CAC alvo")
inputcell(ws, r, 2, 3, INT)
MULT = r
note(ws, r, 3, "Benchmark: >=3:1 e saudavel; 4 a 6:1 e excelente.")
r += 1
label(ws, r, "Churn alternativo para teste de sensibilidade")
inputcell(ws, r, 2, 0.06, PCT)
CHURN_ALT = r
note(ws, r, 3, "Micro-SaaS de ticket baixo para criadores costuma ficar entre 4% e 8%.")
r += 1
label(ws, r, "Fator R: folha minima exigida sobre a receita")
inputcell(ws, r, 2, 0.28, PCT)
FATORR = r
note(ws, r, 3, "Abaixo disso a empresa cai no Anexo V e a margem encolhe.")

TICKET, CHURN = 5, 6
GW_PCT, GW_FIX, SIMPLES, SIMPLES_V, INFRA = 9, 10, 11, 12, 13

# ===================== 2. UNIT ECONOMICS =====================
ws2 = wb.create_sheet("Unit Economics")
title(ws2, "Unit economics por cliente",
      "Calculado a partir da aba Premissas. Verde indica valor vindo de outra aba.")
ws2.column_dimensions["A"].width = 46
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 62

r = 4
section(ws2, r, "COMPOSIÇÃO DA MARGEM")
r += 1
label(ws2, r, "Receita mensal por cliente")
formulacell(ws2, r, 2, f"=Premissas!B{TICKET}", CUR, link=True)
REC = r
r += 1
label(ws2, r, "Taxa do gateway")
formulacell(ws2, r, 2, f"=-(Premissas!B{TICKET}*Premissas!B{GW_PCT}+Premissas!B{GW_FIX})", CUR)
r += 1
label(ws2, r, "Simples Nacional")
formulacell(ws2, r, 2, f"=-(Premissas!B{TICKET}*Premissas!B{SIMPLES})", CUR)
r += 1
label(ws2, r, "Infraestrutura variável")
formulacell(ws2, r, 2, f"=-Premissas!B{INFRA}", CUR)
LASTCOST = r
r += 1
label(ws2, r, "Margem de contribuição por cliente").font = Font(name=FONT, size=10, bold=True)
formulacell(ws2, r, 2, f"=SUM(B{REC}:B{LASTCOST})", CUR, bold=True)
MARGEM = r
r += 1
label(ws2, r, "Margem de contribuição (%)")
formulacell(ws2, r, 2, f"=B{MARGEM}/B{REC}", PCT)

r += 2
section(ws2, r, "VALOR DO CLIENTE")
r += 1
label(ws2, r, "Tempo médio de permanência (meses)")
formulacell(ws2, r, 2, f"=1/Premissas!B{CHURN}", '#,##0.0')
note(ws2, r, 3, "Inverso do churn mensal.")
r += 1
label(ws2, r, "LTV — valor total do cliente").font = Font(name=FONT, size=10, bold=True)
formulacell(ws2, r, 2, f"=B{MARGEM}/Premissas!B{CHURN}", CUR0, bold=True)
LTV = r
r += 1
label(ws2, r, "CAC máximo sustentável").font = Font(name=FONT, size=10, bold=True)
formulacell(ws2, r, 2, f"=B{LTV}/Premissas!B{MULT}", CUR0, bold=True)
note(ws2, r, 3, "Teto de gasto para adquirir um cliente mantendo o multiplo alvo.")
CACMAX = r

r += 2
section(ws2, r, "TESTE DE SENSIBILIDADE")
r += 1
label(ws2, r, "Margem SEM Fator R (Anexo V)")
formulacell(ws2, r, 2, f"=Premissas!B{TICKET}-(Premissas!B{TICKET}*Premissas!B{GW_PCT}+Premissas!B{GW_FIX})-(Premissas!B{TICKET}*Premissas!B{SIMPLES_V})-Premissas!B{INFRA}", CUR)
MARGEM_V = r
note(ws2, r, 3, "Cenario em que a folha nao atinge 28% da receita.")
r += 1
label(ws2, r, "LTV com churn alternativo")
formulacell(ws2, r, 2, f"=B{MARGEM}/Premissas!B{CHURN_ALT}", CUR0)
LTV_ALT = r
r += 1
label(ws2, r, "CAC máximo com churn alternativo")
formulacell(ws2, r, 2, f"=B{LTV_ALT}/Premissas!B{MULT}", CUR0)
note(ws2, r, 3, "Se o churn real for o do segmento de criadores, o teto de CAC cai pela metade.")
r += 1
label(ws2, r, "Receita máxima que o pró-labore atual sustenta no Fator R")
formulacell(ws2, r, 2, f"=Premissas!B{PROL_TOTAL}/Premissas!B{FATORR}", CUR0)
note(ws2, r, 3, "Acima desta receita, e preciso aumentar a folha ou perder o Anexo III.")

# ===================== 3. BREAK-EVEN =====================
ws3 = wb.create_sheet("Break-even")
title(ws3, "Ponto de equilíbrio",
      "Quantos clientes pagantes são necessários em cada nível de compromisso.")
ws3.column_dimensions["A"].width = 46
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 62

r = 4
section(ws3, r, "BASE DE CÁLCULO")
r += 1
label(ws3, r, "Margem de contribuição por cliente")
formulacell(ws3, r, 2, f"='Unit Economics'!B{MARGEM}", CUR, link=True)
BE_MARGEM = r
r += 1
label(ws3, r, "Custo fixo mensal")
formulacell(ws3, r, 2, f"=Premissas!B{FIXED_TOTAL}", CUR0, link=True)
BE_FIXO = r
r += 1
label(ws3, r, "Pró-labore total")
formulacell(ws3, r, 2, f"=Premissas!B{PROL_TOTAL}", CUR0, link=True)
BE_PROL = r

r += 2
section(ws3, r, "CLIENTES NECESSÁRIOS")
r += 1
label(ws3, r, "Para cobrir o custo operacional (COM Fator R)").font = Font(name=FONT, size=10, bold=True)
formulacell(ws3, r, 2, f"=ROUNDUP(B{BE_FIXO}/B{BE_MARGEM},0)", INT, bold=True)
note(ws3, r, 3, "So valido se houver folha >=28% da receita. Sem pro-labore, use a linha da secao Sensibilidade.")
r += 1
label(ws3, r, "Para cobrir custo operacional + pró-labore").font = Font(name=FONT, size=10, bold=True)
formulacell(ws3, r, 2, f"=ROUNDUP((B{BE_FIXO}+B{BE_PROL})/B{BE_MARGEM},0)", INT, bold=True)
note(ws3, r, 3, "A partir daqui o negócio remunera os sócios.")
r += 1
BE_OP = r - 1
label(ws3, r, "Receita mensal no ponto de equilíbrio operacional")
formulacell(ws3, r, 2, f"=B{BE_OP-1}*Premissas!B{TICKET}", CUR0)

r += 2
section(ws3, r, "SENSIBILIDADE")
r += 1
label(ws3, r, "Para cobrir o custo operacional SEM pró-labore (Anexo V)").font = Font(name=FONT, size=10, bold=True)
label(ws3, r, "Para cobrir o custo operacional SEM pró-labore (Anexo V)")
formulacell(ws3, r, 2, f"=ROUNDUP(B{BE_FIXO}/'Unit Economics'!B{MARGEM_V},0)", INT)
note(ws3, r, 3, "Sem folha nao ha Fator R. Este e o numero honesto para o cenario sem pro-labore.")
r += 1
label(ws3, r, "Clientes para custo + pró-labore, com churn alternativo")
formulacell(ws3, r, 2, f"=ROUNDUP((B{BE_FIXO}+B{BE_PROL})/B{BE_MARGEM},0)", INT)
note(ws3, r, 3, "O break-even nao muda com churn; o que muda e o tempo para chegar la.")

# ===================== 4. CENÁRIOS =====================
ws4 = wb.create_sheet("Cenários")
title(ws4, "Projeção de 12 meses após o lançamento",
      "Três cenários de aquisição. Resultado é operacional, antes de pró-labore.")

ws4.column_dimensions["A"].width = 10
for col in range(2, 11):
    ws4.column_dimensions[get_column_letter(col)].width = 15

hdr_row = 4
groups = [("CONSERVADOR", 2), ("BASE", 5), ("OTIMISTA", 8)]
for name, col in groups:
    c = ws4.cell(row=hdr_row - 1, column=col, value=name)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HEADFILL
    c.alignment = Alignment(horizontal="center")
    ws4.merge_cells(start_row=hdr_row - 1, start_column=col, end_row=hdr_row - 1, end_column=col + 2)

headers = ["Mês", "Clientes", "Receita", "Resultado",
           "Clientes", "Receita", "Resultado",
           "Clientes", "Receita", "Resultado"]
for i, h in enumerate(headers, start=1):
    c = ws4.cell(row=hdr_row, column=i, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color=DARK)
    c.fill = SOFTFILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

scen_refs = {2: CEN_CONS, 5: CEN_BASE, 8: CEN_OTIM}

first = hdr_row + 1
for m in range(1, 13):
    row = hdr_row + m
    c = ws4.cell(row=row, column=1, value=m)
    c.font = Font(name=FONT, size=10, color=BLACK)
    c.alignment = Alignment(horizontal="center")
    for col, prem_row in scen_refs.items():
        L = get_column_letter(col)
        if m == 1:
            f = f"=Premissas!$B${prem_row}"
        else:
            f = f"={L}{row-1}*(1-Premissas!$B${CHURN})+Premissas!$B${prem_row}"
        formulacell(ws4, row, col, f, INT)
        formulacell(ws4, row, col + 1, f"={L}{row}*Premissas!$B${TICKET}", CUR0)
        formulacell(ws4, row, col + 2,
                    f"={L}{row}*'Unit Economics'!$B${MARGEM}-Premissas!$B${FIXED_TOTAL}", CUR0)

r = hdr_row + 14
note(ws4, r, 1, "Clientes = base do mês anterior menos churn, mais novos clientes do cenário.")
r += 1
note(ws4, r, 1, "Resultado = margem de contribuição total menos custo fixo. Não inclui pró-labore.")
r += 1
note(ws4, r, 1, "As taxas de aquisição são premissa, não previsão. Ajuste na aba Premissas assim que houver dado real.")

# ===================== 5. CUSTO DE CONSTRUÇÃO =====================
ws5 = wb.create_sheet("Construção 90 dias")
title(ws5, "Desembolso do período de construção",
      "Custo em caixa dos 90 dias até o lançamento. Não inclui remuneração dos sócios.")
ws5.column_dimensions["A"].width = 46
ws5.column_dimensions["B"].width = 16
ws5.column_dimensions["C"].width = 16
ws5.column_dimensions["D"].width = 56

r = 4
for i, h in enumerate(["Item", "Mínimo", "Máximo", "Observação"], start=1):
    c = ws5.cell(row=r, column=i, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color=DARK)
    c.fill = SOFTFILL
    c.border = BORDER

itens = [
    ("Abertura de CNPJ e contabilidade (3 meses)", 2400, 2400, "Abertura ~R$1.500 + R$300/mês."),
    ("Assessoria jurídica", 3000, 6000, "Acordo de quotistas e parecer sobre canais. Faixa ampla."),
    ("Infraestrutura durante o desenvolvimento", 180, 180, "VPS pequena por 3 meses."),
    ("Domínio, e-mail e ferramentas", 450, 450, "Maior parte em free tier."),
    ("Reserva para imprevistos", 1000, 1000, "Estimativa própria."),
]
start = r + 1
for i, (nome, mn, mx, obs) in enumerate(itens):
    row = start + i
    label(ws5, row, nome)
    inputcell(ws5, row, 2, mn, CUR0)
    inputcell(ws5, row, 3, mx, CUR0)
    note(ws5, row, 4, obs)
end = start + len(itens) - 1

row = end + 1
label(ws5, row, "Total").font = Font(name=FONT, size=10, bold=True)
formulacell(ws5, row, 2, f"=SUM(B{start}:B{end})", CUR0, bold=True)
formulacell(ws5, row, 3, f"=SUM(C{start}:C{end})", CUR0, bold=True)

row += 2
section(ws5, row, "COMPARAÇÃO")
row += 1
label(ws5, row, "Meses de custo fixo cobertos pelo desembolso máximo")
formulacell(ws5, row, 2, f"=C{end+1}/Premissas!B{FIXED_TOTAL}", '#,##0.0')
note(ws5, row, 4, "Quanto tempo o valor da construção sustentaria a operação já no ar.")

# ===================== LEGENDA =====================
for sheet in (ws, ws2, ws3, ws4, ws5):
    last = sheet.max_row + 3
    c = sheet.cell(row=last, column=1, value="LEGENDA")
    c.font = Font(name=FONT, size=9, bold=True, color="808080")
    c2 = sheet.cell(row=last + 1, column=1,
                    value="Amarelo com texto azul = entrada editável   •   Preto = fórmula   •   Verde = valor de outra aba")
    c2.font = Font(name=FONT, size=9, italic=True, color="808080")

wb.save(OUT)
print("OK:", OUT)
